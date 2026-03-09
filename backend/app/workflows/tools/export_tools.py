"""Export workflow tools — generate Excel reports and other output formats."""

from __future__ import annotations

import io
import os
import uuid

import structlog

from app.workflows.tools.base import BaseTool, ToolInput, ToolOutput

logger = structlog.get_logger(__name__)

EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "exports")


class ExportExcelTool(BaseTool):
    name = "export.excel"
    description = "Generate an Excel file from structured data (e.g. compliance matrix)."
    input_schema = {
        "type": "object",
        "properties": {
            "requirements": {"type": "array", "description": "List of requirement objects"},
            "compliance": {"type": "array", "description": "List of compliance result objects"},
            "template": {"type": "string", "description": "Template name (e.g. 'compliance_matrix')"},
            "filename": {"type": "string", "default": "export.xlsx"},
        },
        "required": ["requirements"],
    }
    output_schema = {
        "type": "object",
        "properties": {
            "file_url": {"type": "string"},
            "file_path": {"type": "string"},
        },
    }

    async def execute(self, tool_input: ToolInput) -> ToolOutput:
        params = tool_input.params
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Compliance Matrix"

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

            template = params.get("template", "compliance_matrix")
            requirements = params.get("requirements", [])
            compliance = params.get("compliance", [])

            if template == "compliance_matrix":
                headers = ["Req ID", "Title", "Description", "Status", "Confidence", "Evidence", "Gaps", "Suggested Action"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Build compliance lookup
                compliance_map = {}
                for c in compliance:
                    if isinstance(c, dict):
                        key = c.get("id") or c.get("query", "")
                        compliance_map[key] = c

                for i, req in enumerate(requirements, 2):
                    if not isinstance(req, dict):
                        continue
                    req_id = req.get("id", f"REQ-{i-1}")
                    comp = compliance_map.get(req_id, {})

                    ws.cell(row=i, column=1, value=req_id)
                    ws.cell(row=i, column=2, value=req.get("title", ""))
                    ws.cell(row=i, column=3, value=req.get("description", ""))
                    ws.cell(row=i, column=4, value=comp.get("status", "Not Checked"))
                    ws.cell(row=i, column=5, value=comp.get("confidence", ""))
                    ws.cell(row=i, column=6, value=comp.get("evidence", ""))
                    ws.cell(row=i, column=7, value=comp.get("gaps", ""))
                    ws.cell(row=i, column=8, value=comp.get("suggested_action", ""))

                    # Color-code status
                    status = comp.get("status", "").lower()
                    status_cell = ws.cell(row=i, column=4)
                    if "met" == status:
                        status_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                    elif "partially" in status:
                        status_cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
                    elif "not met" in status:
                        status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

                # Auto-width columns
                for col in ws.columns:
                    max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

                # Summary sheet
                summary = wb.create_sheet("Summary")
                total = len(requirements)
                met = sum(1 for c in compliance if isinstance(c, dict) and c.get("status", "").lower() == "met")
                partial = sum(1 for c in compliance if isinstance(c, dict) and "partial" in c.get("status", "").lower())
                not_met = sum(1 for c in compliance if isinstance(c, dict) and "not met" in c.get("status", "").lower())
                summary.cell(row=1, column=1, value="Total Requirements").font = Font(bold=True)
                summary.cell(row=1, column=2, value=total)
                summary.cell(row=2, column=1, value="Met")
                summary.cell(row=2, column=2, value=met)
                summary.cell(row=3, column=1, value="Partially Met")
                summary.cell(row=3, column=2, value=partial)
                summary.cell(row=4, column=1, value="Not Met")
                summary.cell(row=4, column=2, value=not_met)
            else:
                # Generic table export
                if requirements:
                    first = requirements[0] if isinstance(requirements[0], dict) else {}
                    headers = list(first.keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                    for i, item in enumerate(requirements, 2):
                        if isinstance(item, dict):
                            for col, key in enumerate(headers, 1):
                                ws.cell(row=i, column=col, value=str(item.get(key, "")))

            # Save
            filename = params.get("filename", "export.xlsx")
            os.makedirs(EXPORTS_DIR, exist_ok=True)
            file_path = os.path.join(EXPORTS_DIR, f"{uuid.uuid4().hex[:8]}_{filename}")
            wb.save(file_path)

            return ToolOutput(
                success=True,
                data={
                    "file_path": file_path,
                    "file_url": f"/api/exports/{os.path.basename(file_path)}",
                    "filename": filename,
                },
            )
        except Exception as e:
            logger.exception("export_excel_failed", error=str(e))
            return ToolOutput(success=False, error=str(e))
